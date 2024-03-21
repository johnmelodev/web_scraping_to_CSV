# 1 - Preciso de uma automação para fazer a cotação em 2 sites na internet
# - kabum
# - https://www.studiopc.com.br
# “memória ram 16gb"
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import *
import openpyxl
import os
import pyautogui
from time import sleep
import schedule
import pyperclip


def run_bot_main():
    def get_price():
        def start_driver():
            chrome_options = Options()

            arguments = ['--lang=en-us',
                         '--window-size=1920,1080', '--incognito']

            for argument in arguments:
                chrome_options.add_argument(argument)

            chrome_options.add_experimental_option('prefs', {
                'download.prompt_for_download': False,
                'profile.default_content_setting_values.notifications': 2,
                'profile.default_content_setting_values.automatic_downloads': 1
            })

            driver = webdriver.Chrome(service=ChromeService(
                ChromeDriverManager().install()), options=chrome_options)
            wait = WebDriverWait(
                driver,
                10,
                poll_frequency=1,
                ignored_exceptions=[
                    NoSuchElementException,
                    ElementNotVisibleException,
                    ElementNotSelectableException,
                ]
            )
            return driver, wait

        driver, wait = start_driver()

        driver.get(
            'https://www.kabum.com.br/busca/memoria-ram-16-gb?page_number=1&page_size=20&facet_filters=&sort=price&variant=catalog')

        ram_memory_1 = wait.until(expected_conditions.visibility_of_all_elements_located(
            (By.XPATH, "//span[@class='sc-b1f5eb03-2 iaiQNF priceCard']")))

        ram_memory_price_1 = float(
            ram_memory_1[0].text.split(' ')[1].replace(',', '.'))

        driver.get('https://www.studiopc.com.br/produtos?q=memoria+ram+16gb')

        ram_memory_2 = wait.until(expected_conditions.visibility_of_any_elements_located(
            (By.XPATH, "//span[@class='price total']")))

        ram_memory_price_2 = float(
            ram_memory_2[0].text.split(' ')[1].replace(',', '.'))

        return ram_memory_price_1, ram_memory_price_2

    def generate_profit_margin_spreadsheet():
        # 2 - Depois disso jogar em uma planilha de excel e calcular a margem(lucro)
        ram_memory_1, ram_memory_2 = get_price()

        cost = 200
        site_1 = 'https://www.studiopc.com.br'
        site_2 = 'https://www.kabum.com.br'

        workbook = openpyxl.Workbook()
        del workbook['Sheet']
        workbook.create_sheet('profit_margin')
        sheet_profit_margin = workbook['profit_margin']
        sheet_profit_margin.append(['Site', 'Cost', 'Price', 'Profit'])
        sheet_profit_margin.append(
            [site_1, cost, ram_memory_1, ram_memory_1 - cost])
        sheet_profit_margin.append(
            [site_2, cost, ram_memory_2, ram_memory_2 - cost])
        workbook.save('margin of profit.xlsx')

        workbook_profit_margin = openpyxl.load_workbook(
            'margin of profit.xlsx')
        sheet_profit_margin = workbook_profit_margin['profit_margin']
        margin_of_profit = ''
        for line in sheet_profit_margin.iter_rows(min_row=1):
            margin_of_profit += f'{line[0].value},{
                line[1].value},{line[2].value},{line[3].value}{os.linesep}'

        with open('profit_margin.txt', 'w', newline='', encoding='utf-8') as file:
            file.write(margin_of_profit)

        return margin_of_profit

    def sent_profit_margin_to_whatsapp():
        # 3 - enviar em um grupo do whatsapp
        margin_of_profit = generate_profit_margin_spreadsheet()

        pyautogui.keyDown('command')
        sleep(0.5)
        pyautogui.press('space')
        pyautogui.keyUp('command')
        sleep(2)
        pyautogui.write('whatsapp')
        sleep(2)
        pyautogui.hotkey('enter')
        sleep(3)
        pyautogui.doubleClick(390, 135)
        sleep(2)

        # name of whatsapp group or contact
        pyautogui.write('my code test')
        sleep(2)
        pyautogui.move(0, 120)
        sleep(1)
        pyautogui.click()
        sleep(2)
        pyautogui.doubleClick(788, 808)
        sleep(2)
        pyperclip.copy(margin_of_profit)
        pyautogui.hotkey('command', 'v')
        sleep(2)
        pyautogui.hotkey('enter')
    sent_profit_margin_to_whatsapp()


# 4 - todos os dias as 06:00 da manhã

schedule.every().day.at("18:40").do(run_bot_main)
run_bot_main()
